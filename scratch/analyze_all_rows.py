import openpyxl

wb_path = '/Applications/Arenga/vscode/breakdown_app/master rekap 2026_Bom.xlsx'
wb = openpyxl.load_workbook(wb_path, data_only=False)

# Let's inspect the sheets and defined names
print("Workbook sheets:", wb.sheetnames)
print(f"Total defined names: {len(wb.defined_names)}")

# Let's list some defined names
defined_names_info = []
for name, dn in wb.defined_names.items():
    defined_names_info.append((name, dn.value))

# Search for relevant defined names
relevant_dn = ['tabprt', 'tbl_edg', 'std_bhn_prt', 'tabdf', 'tabkab']
print("\n--- Relevant Defined Names ---")
for r_name in relevant_dn:
    found = False
    for name, val in defined_names_info:
        if name.lower() == r_name.lower():
            print(f"  {name}: {val}")
            found = True
    if not found:
        # search substring
        for name, val in defined_names_info:
            if r_name.lower() in name.lower():
                print(f"  {name} (partial match): {val}")

# Let's see defined names related to Stock, Spek, or Data Validation
print("\n--- Defined Names referring to Stock, Spek, Data Validation ---")
for name, val in defined_names_info:
    val_str = str(val)
    if 'Stock' in val_str or 'Spek' in val_str or 'Data Validation' in val_str:
        print(f"  {name}: {val}")

# Now let's analyze rows 14 to 198 in Breakdown
sheet = wb['Breakdown']
row_types = {}
column_formulas = {} # column -> set of formulas

cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY', 'BZ', 'CA', 'CB', 'CC', 'CD', 'CE', 'CF']

print("\nScanning Breakdown rows 14-198...")
for r in range(14, 199):
    r_type = sheet[f'C{r}'].value
    if r_type is None:
        continue
    row_types[r_type] = row_types.get(r_type, 0) + 1
    
    # Check formulas
    for col in cols:
        val = sheet[f'{col}{r}'].value
        if val is not None and str(val).startswith('='):
            # Normalize formula by removing row number
            normalized = str(val)
            # A simple normalization: replace current row number with 'n'
            # (e.g., A14 -> An, B14 -> Bn, Spek!$E$10 stays Spek!$E$10)
            # Let's replace the string representation of r
            normalized = normalized.replace(str(r), 'n')
            if col not in column_formulas:
                column_formulas[col] = {}
            column_formulas[col][normalized] = column_formulas[col].get(normalized, 0) + 1

print("\nRow type counts in rows 14-198:")
for t, count in row_types.items():
    print(f"  {t}: {count}")

print("\nUnique formulas per column in rows 14-198:")
for col in sorted(column_formulas.keys()):
    print(f"  Col {col}:")
    for formula, count in column_formulas[col].items():
        print(f"    ({count} rows) {formula}")

# Let's read Spek sheet to see what it contains
spek_sheet = wb['Spek']
print("\n--- Spek Sheet Contents (rows 1-30, cols A-G) ---")
for r in range(1, 31):
    row_vals = [spek_sheet.cell(r, c).value for c in range(1, 8)]
    if any(v is not None for v in row_vals):
        print(f"  Row {r}: {row_vals}")

# Let's read Data Validation sheet to see what it contains
dv_sheet = wb['Data Validation']
print("\n--- Data Validation Sheet Contents (rows 1-30, cols A-H) ---")
for r in range(1, 31):
    row_vals = [dv_sheet.cell(r, c).value for c in range(1, 9)]
    if any(v is not None for v in row_vals):
        print(f"  Row {r}: {row_vals}")
