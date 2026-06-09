import openpyxl

wb_path = "/Applications/Arenga/vscode/breakdown_app/master rekap 2026_Bom.xlsx"
print("Loading workbook with formulas...")
wb_formula = openpyxl.load_workbook(wb_path, data_only=False)

print("Loading workbook with values...")
wb_value = openpyxl.load_workbook(wb_path, data_only=True)

# 1. Inspect Spek sheet around D43
if 'Spek' in wb_formula.sheetnames:
    sheet_formula = wb_formula['Spek']
    sheet_value = wb_value['Spek']
    print("\n=== Spek Sheet Column D (Rows 35 to 55) ===")
    for r in range(35, 56):
        cell_f = sheet_formula.cell(row=r, column=4) # Column D
        cell_v = sheet_value.cell(row=r, column=4)
        label_cell = sheet_value.cell(row=r, column=3) # Column C
        print(f"Row {r} | C (Label): {label_cell.value} | D (Formula): {cell_f.value} | D (Value): {cell_v.value}")
else:
    print("Spek sheet not found!")

# 2. Inspect Data Validation sheet around rows 9 to 25
if 'Data Validation' in wb_formula.sheetnames:
    val_formula = wb_formula['Data Validation']
    val_value = wb_value['Data Validation']
    print("\n=== Data Validation Column D (Rows 9 to 25) ===")
    for r in range(9, 26):
        # Column A, B, C, D
        col_A_f = val_formula.cell(row=r, column=1).value
        col_B_f = val_formula.cell(row=r, column=2).value
        col_C_f = val_formula.cell(row=r, column=3).value
        col_D_f = val_formula.cell(row=r, column=4).value
        
        col_A_v = val_value.cell(row=r, column=1).value
        col_B_v = val_value.cell(row=r, column=2).value
        col_C_v = val_value.cell(row=r, column=3).value
        col_D_v = val_value.cell(row=r, column=4).value
        
        print(f"Row {r} | A: {col_A_v} ({col_A_f}) | B: {col_B_v} ({col_B_f}) | C: {col_C_v} ({col_C_f}) | D: {col_D_v} ({col_D_f})")
else:
    print("Data Validation sheet not found!")

wb_formula.close()
wb_value.close()
