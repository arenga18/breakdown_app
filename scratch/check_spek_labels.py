import openpyxl

wb_path = "/Applications/Arenga/vscode/breakdown_app/master rekap 2026_Bom.xlsx"
print("Loading workbook...")
wb = openpyxl.load_workbook(wb_path, data_only=True)

if 'Spek' in wb.sheetnames:
    sheet = wb['Spek']
    print("\n=== Spek Sheet Rows 35 to 70 ===")
    for r in range(35, 71):
        # Print column A, B, C, D, E, F
        row_vals = []
        for c in range(1, 8):
            val = sheet.cell(row=r, column=c).value
            row_vals.append(f"Col{c}: {val}")
        print(f"Row {r:2d} | " + " | ".join(row_vals))
else:
    print("Spek sheet not found!")

wb.close()
