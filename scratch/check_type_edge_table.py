import openpyxl

wb_path = "/Applications/Arenga/vscode/breakdown_app/master rekap 2026_Bom.xlsx"
print("Loading workbook with formulas...")
wb_formula = openpyxl.load_workbook(wb_path, data_only=False)

print("Loading workbook with values...")
wb_value = openpyxl.load_workbook(wb_path, data_only=True)

if 'Data Validation' in wb_formula.sheetnames:
    sheet_f = wb_formula['Data Validation']
    sheet_v = wb_value['Data Validation']
    
    print("\n=== Data Validation Sheet Rows 515 to 545 ===")
    for r in range(515, 546):
        row_str = []
        for c in range(1, 10):
            val_v = sheet_v.cell(row=r, column=c).value
            val_f = sheet_f.cell(row=r, column=c).value
            if val_v is not None or val_f is not None:
                row_str.append(f"Col{c}: {val_v} ({val_f})")
        if row_str:
            print(f"Row {r} | " + " | ".join(row_str))
else:
    print("Data Validation sheet not found!")

wb_formula.close()
wb_value.close()
