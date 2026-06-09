import openpyxl

wb_path = "/Applications/Arenga/vscode/breakdown_app/master rekap 2026_Bom.xlsx"
print("Loading workbook with formulas...")
wb_formula = openpyxl.load_workbook(wb_path, data_only=False)

print("Loading workbook with values...")
wb_value = openpyxl.load_workbook(wb_path, data_only=True)

if 'Data Validation' in wb_formula.sheetnames:
    sheet_f = wb_formula['Data Validation']
    sheet_v = wb_value['Data Validation']
    
    print("Searching for 'Type Finished' or 'ctf'...")
    matches = []
    for r in range(1, 1000):
        for c in range(1, 50):
            val = sheet_v.cell(row=r, column=c).value
            if val == "Type Finished" or val == "ctf" or (isinstance(val, str) and "Type Finished" in val):
                matches.append((r, c, val))
    
    print(f"Found matches: {matches}")
    
    # For each match, let's print the block around it (25 rows down, 8 columns wide)
    for r_idx, c_idx, val in matches:
        print(f"\n=== Inspection around Row {r_idx}, Col {c_idx} ({val}) ===")
        for r in range(r_idx, r_idx + 25):
            row_str = []
            for c in range(c_idx - 2, c_idx + 8):
                if c < 1: continue
                val_v = sheet_v.cell(row=r, column=c).value
                val_f = sheet_f.cell(row=r, column=c).value
                if val_v is not None or val_f is not None:
                    row_str.append(f"Col{c}: {val_v} ({val_f})")
            if row_str:
                print(f"Row {r} | " + " | ".join(row_str))
else:
    print("Data Validation sheet not found!")

wb_formula.close()
wb_value.close()
