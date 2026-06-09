import openpyxl
import json

wb = openpyxl.load_workbook('/Applications/Arenga/vscode/breakdown_app/master rekap 2026_Bom.xlsx', read_only=True, data_only=False)
sheet = wb['Bom']

rows = []
for r in range(13, 250): # Let's read from row 13 to 250
    row_data = {
        "row_idx": r,
        "line_no": sheet.cell(row=r, column=1).value or "",
        "id_barang": sheet.cell(row=r, column=2).value or "",
        "nama_barang": sheet.cell(row=r, column=3).value or "",
        "jml_formula": sheet.cell(row=r, column=4).value or "",
        "satuan": sheet.cell(row=r, column=5).value or "",
        "keterangan": sheet.cell(row=r, column=6).value or ""
    }
    
    # Check if there is any data in the row
    is_empty = all(not v for k, v in row_data.items() if k != "row_idx")
    if not is_empty:
        rows.append(row_data)
    else:
        # If it's a completely empty row, let's see if we should preserve it as spacing
        rows.append({
            "row_idx": r,
            "line_no": "",
            "id_barang": "",
            "nama_barang": "",
            "jml_formula": "",
            "satuan": "",
            "keterangan": ""
        })

# Save to a json file
with open('/Applications/Arenga/vscode/breakdown_app/scratch/bom_setting_full.json', 'w') as f:
    json.dump(rows, f, indent=2)

print(f"Dumped {len(rows)} rows to scratch/bom_setting_full.json")
# Find the last non-empty row to prune the list
last_non_empty_idx = 0
for i, r in enumerate(rows):
    is_empty = all(not v for k, v in r.items() if k != "row_idx")
    if not is_empty:
        last_non_empty_idx = i

print(f"Last non-empty row index in our list: {last_non_empty_idx + 13} (list index {last_non_empty_idx})")
