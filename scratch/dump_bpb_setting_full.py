import openpyxl
import json

wb = openpyxl.load_workbook('/Applications/Arenga/vscode/breakdown_app/master rekap 2026_Bom.xlsx', read_only=True, data_only=False)
sheet = wb['BPB Setting']

rows = []
for r in range(13, 142): # Row 13 to 141
    row_no = sheet.cell(row=r, column=1).value
    id_barang = sheet.cell(row=r, column=2).value
    nama_barang = sheet.cell(row=r, column=3).value
    jml_formula = sheet.cell(row=r, column=4).value
    satuan = sheet.cell(row=r, column=5).value
    keterangan = sheet.cell(row=r, column=6).value
    
    rows.append({
        "row_idx": r,
        "line_no": f"{row_no}" if row_no is not None else "",
        "id_barang": f"{id_barang}" if id_barang is not None else "",
        "nama_barang": f"{nama_barang}" if nama_barang is not None else "",
        "jml_formula": f"{jml_formula}" if jml_formula is not None else "",
        "satuan": f"{satuan}" if satuan is not None else "",
        "keterangan": f"{keterangan}" if keterangan is not None else ""
    })

with open('/Applications/Arenga/vscode/breakdown_app/scratch/bpb_setting_full.json', 'w') as f:
    json.dump(rows, f, indent=2)

print(f"Dumped {len(rows)} rows to scratch/bpb_setting_full.json")
