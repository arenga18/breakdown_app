import openpyxl
import json

wb_path = "/Applications/Arenga/vscode/breakdown_app/master rekap 2026_Bom.xlsx"

print("Loading workbook for formulas (read_only)...", flush=True)
wb_form = openpyxl.load_workbook(wb_path, read_only=True, data_only=False)
sheet_form = wb_form['Breakdown']

print("Loading workbook for values (read_only)...", flush=True)
wb_val = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
sheet_val = wb_val['Breakdown']

print("Extracting rows 14 to 198...", flush=True)

rows_form = list(sheet_form.iter_rows(min_row=11, max_row=198))
rows_val = list(sheet_val.iter_rows(min_row=11, max_row=198))

headers = [cell.value for cell in rows_val[0]]

# Map of column letter to header
col_map = {}
for idx, h in enumerate(headers):
    letter = openpyxl.utils.get_column_letter(idx+1)
    col_map[letter] = h or f"Col_{letter}"

results = []

for r in range(14, 199):
    idx = r - 11
    if idx >= len(rows_form):
        break
    cells_form = rows_form[idx]
    cells_val = rows_val[idx]
    
    # Check if row is empty
    vals = [c.value for c in cells_val]
    if all(v is None for v in vals):
        continue
        
    row_data = {
        'row_number': r,
        'columns': {}
    }
    
    for c_idx in range(len(headers)):
        letter = openpyxl.utils.get_column_letter(c_idx+1)
        val_cell = cells_val[c_idx].value
        form_cell = cells_form[c_idx].value
        
        # We only record non-empty columns to keep it concise
        if val_cell is not None or form_cell is not None:
            row_data['columns'][letter] = {
                'header': col_map.get(letter, f"Col_{letter}"),
                'value': val_cell,
                'formula': form_cell if str(form_cell).startswith('=') else None
            }
    results.append(row_data)

# Extract defined names in workbook
defined_names_dict = {}
for name, dn in wb_form.defined_names.items():
    defined_names_dict[name] = dn.value

output_data = {
    'defined_names': defined_names_dict,
    'rows': results
}

with open("/Applications/Arenga/vscode/breakdown_app/scratch/breakdown_rows_14_198.json", "w") as f:
    json.dump(output_data, f, indent=2)

print("Extraction completed! Saved to scratch/breakdown_rows_14_198.json", flush=True)

wb_form.close()
wb_val.close()
