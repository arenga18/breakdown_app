import openpyxl

wb = openpyxl.load_workbook('/Applications/Arenga/vscode/breakdown_app/master rekap 2026_Bom.xlsx', read_only=True, data_only=False)
sheet = wb['Bom']

print("Row index, Column C (Name), Column D (Jml Formula), Column K (Formula/Value):")
for r in range(13, 150):
    name = sheet.cell(row=r, column=3).value
    col_d = sheet.cell(row=r, column=4).value
    col_k = sheet.cell(row=r, column=11).value
    if name or col_d or col_k:
        print(f"Row {r:03d} | C: {name} | D: {col_d} | K: {col_k}")
