import openpyxl

wb = openpyxl.load_workbook('/Applications/Arenga/vscode/breakdown_app/master rekap 2026_Bom.xlsx', read_only=True, data_only=False)
sheet = wb['Bom']

# Print columns A to P for rows 20 to 35
for r in range(20, 36):
    vals = [sheet.cell(row=r, column=c).value for c in range(1, 17)]
    if any(v is not None for v in vals):
        # Format the values nicely
        formatted_vals = [f"Col {openpyxl.utils.get_column_letter(c)}: {v}" for c, v in enumerate(vals, 1) if v is not None]
        print(f"Row {r:02d}: {', '.join(formatted_vals)}")
