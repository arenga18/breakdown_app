import openpyxl

wb = openpyxl.load_workbook('/Applications/Arenga/vscode/breakdown_app/master rekap 2026_Bom.xlsx', read_only=True, data_only=False)
sheet = wb['Bom']

print("Row index, Column C (Name), Column D (Jml Formula):")
for r in range(140, 210):
    name = sheet.cell(row=r, column=3).value
    col_d = sheet.cell(row=r, column=4).value
    if name or col_d:
        print(f"Row {r:03d} | C: {name} | D: {col_d}")
