import openpyxl
import json

wb = openpyxl.load_workbook('/Applications/Arenga/vscode/breakdown_app/master rekap 2026_Bom.xlsx', read_only=True, data_only=False)
sheet = wb['Bom']

print(f"Max row in Bom: {sheet.max_row}")
print(f"Max col in Bom: {sheet.max_column}")

rows = []
for r in range(1, 150):
    vals = [sheet.cell(row=r, column=c).value for c in range(1, 8)]
    # Check if row is not empty
    if any(v is not None for v in vals):
        rows.append((r, vals))

print("Rows in Bom (first 100 non-empty):")
for r, vals in rows[:100]:
    print(f"Row {r:03d}: {vals}")
