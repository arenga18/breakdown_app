import openpyxl

wb = openpyxl.load_workbook('/Applications/Arenga/vscode/breakdown_app/master rekap 2026_Bom.xlsx', data_only=False)
sheet = wb['BPB Setting']

print("Rows in BPB Setting:")
for r in range(1, sheet.max_row + 1):
    vals = [sheet.cell(row=r, column=c).value for c in range(1, 8)]
    if any(vals):
        # Print row index and non-empty values
        print(f"Row {r:03d}: {[v if v is not None else '' for v in vals]}")
