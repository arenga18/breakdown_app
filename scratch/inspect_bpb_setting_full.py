import openpyxl

wb = openpyxl.load_workbook('/Applications/Arenga/vscode/breakdown_app/master rekap 2026_Bom.xlsx', data_only=False)
sheet = wb['BPB Setting']

print(f"Reading all rows of sheet 'BPB Setting':")
for r in range(1, sheet.max_row + 1):
    row_vals = []
    for c in range(1, 8): # read columns A to G
        cell = sheet.cell(row=r, column=c)
        val = cell.value
        row_vals.append(f"{val}" if val is not None else "")
    if any(row_vals):
        print(f"Row {r:03d}: {row_vals}")
