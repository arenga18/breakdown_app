import openpyxl

wb = openpyxl.load_workbook('/Applications/Arenga/vscode/breakdown_app/master rekap 2026_Bom.xlsx', data_only=False)
sheet = wb['Breakdown']

# Read column headers at row 22 and first data row at row 23
headers = {}
row22 = []
row23 = []
for c in range(1, sheet.max_column + 1):
    h = sheet.cell(row=22, column=c).value
    v = sheet.cell(row=23, column=c).value
    col_letter = openpyxl.utils.get_column_letter(c)
    headers[col_letter] = (h, v)

# Print some columns that are relevant
for col, (h, v) in headers.items():
    if h or v:
        print(f"Col {col} ({h}): {v}")
