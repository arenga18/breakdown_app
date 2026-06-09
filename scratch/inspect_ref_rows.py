import openpyxl

wb_path = "/Applications/Arenga/vscode/breakdown_app/master rekap 2026_Bom.xlsx"
wb_form = openpyxl.load_workbook(wb_path, read_only=True, data_only=False)
sheet_form = wb_form['Breakdown']

wb_val = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
sheet_val = wb_val['Breakdown']

rows_form = list(sheet_form.iter_rows(min_row=169, max_row=170))
rows_val = list(sheet_val.iter_rows(min_row=169, max_row=170))

headers = [cell.value for cell in list(wb_val['Breakdown'].iter_rows(min_row=11, max_row=11))[0]]

print("=== INSPECTING ROW 169 (Ref) ===")
cells_form = rows_form[0]
cells_val = rows_val[0]

for c_idx in range(len(headers)):
    letter = openpyxl.utils.get_column_letter(c_idx+1)
    val = cells_val[c_idx].value
    form = cells_form[c_idx].value
    header = headers[c_idx] or f"Col_{letter}"
    
    if val is not None or form is not None:
        is_formula = str(form).startswith('=')
        formula_str = f" | Formula: {form}" if is_formula else ""
        print(f"Col {letter} ({header}): Value = {val}{formula_str}")

wb_form.close()
wb_val.close()
