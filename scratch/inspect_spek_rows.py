import openpyxl

wb = openpyxl.load_workbook("master rekap 2026_Bom.xlsx", read_only=True)
sheet = wb["Spek"]

print("SPEK SHEET ROWS:")
for r in range(1, 70):
    val_a = sheet.cell(row=r, column=1).value
    val_b = sheet.cell(row=r, column=2).value
    val_c = sheet.cell(row=r, column=3).value
    val_d = sheet.cell(row=r, column=4).value
    if val_a or val_b or val_c or val_d:
        print(f"Row {r:02d} | A: {val_a} | B: {val_b} | C: {val_c} | D: {val_d}")
