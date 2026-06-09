import openpyxl

wb_path = "/Applications/Arenga/vscode/breakdown_app/master rekap 2026_Bom.xlsx"
wb_val = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)

if 'Spek' not in wb_val.sheetnames:
    print("Sheet 'Spek' not found!")
    wb_val.close()
    exit(1)

sheet = wb_val['Spek']

# Read rows 170 to 200
print("=== Spek Sheet Rows 170 to 200 ===")
# We will print all columns that have non-empty headers or values
# Let's first read row 175-177 headers
headers = [sheet.cell(row=r, column=col).value for r in range(170, 178) for col in range(1, 20)]

for r in range(170, 201):
    row_vals = [sheet.cell(row=r, column=col).value for col in range(1, 20)]
    if any(v is not None for v in row_vals):
        # Format values
        vals_str = " | ".join([f"Col{c+1}: {v}" for c, v in enumerate(row_vals) if v is not None])
        print(f"Row {r}: {vals_str}")

wb_val.close()
