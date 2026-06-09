import openpyxl
import sys

wb_path = "/Applications/Arenga/vscode/breakdown_app/master rekap 2026_Bom.xlsx"

print("Loading workbook to check sheets...", flush=True)
wb_sheets = openpyxl.load_workbook(wb_path, read_only=True)
print("Sheets:", wb_sheets.sheetnames, flush=True)
wb_sheets.close()

print("Loading workbook with formulas...", flush=True)
wb_form = openpyxl.load_workbook(wb_path, data_only=False)
print("Workbook loaded!", flush=True)

# Print some defined names
print("\n--- Defined Names (first 40) ---", flush=True)
count = 0
for name, dn in wb_form.defined_names.items():
    print(f"DefinedName: {name} -> {dn.value}")
    count += 1
    if count >= 40:
        break

# Breakdown sheet formulas
sheet_name = 'Breakdown'
if sheet_name not in wb_form.sheetnames:
    print(f"Sheet '{sheet_name}' not found. Available: {wb_form.sheetnames}", flush=True)
    sys.exit(1)

sheet = wb_form[sheet_name]

# Let's inspect headers in row 11
headers = [sheet.cell(row=11, column=col).value for col in range(1, 100)]
print("\n--- Headers Row 11 ---", flush=True)
for idx, h in enumerate(headers):
    if h:
        print(f"Col {idx+1} ({openpyxl.utils.get_column_letter(idx+1)}): {h}")

# Let's print rows 14 to 30 as a starting point to see what's there
print("\n--- Rows 14 to 30 (Formulas) ---", flush=True)
for r in range(14, 31):
    row_vals = []
    has_content = False
    for col in range(1, 40):
        val = sheet.cell(row=r, column=col).value
        if val is not None:
            has_content = True
        row_vals.append(val)
    if has_content:
        # Print column C (type), H (component name), J (P), L (L), N (T), R (Bahan), V (Luar), X (Dalam), Z (P1), AD (P1 edg)
        col_c = row_vals[2] # Type
        col_h = row_vals[7] # Komponen
        col_j = row_vals[9] # P
        col_l = row_vals[11] # L
        col_r = row_vals[17] # Bahan
        col_v = row_vals[21] # Luar
        col_x = row_vals[23] # Dalam
        print(f"Row {r} | Type: {col_c} | Komp: {col_h} | P: {col_j} | L: {col_l} | Bahan: {col_r} | Luar: {col_v} | Dalam: {col_x}", flush=True)
