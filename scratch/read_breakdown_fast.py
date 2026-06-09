import openpyxl

wb_path = "/Applications/Arenga/vscode/breakdown_app/master rekap 2026_Bom.xlsx"

print("Loading workbook for formulas (read_only)...", flush=True)
wb_form = openpyxl.load_workbook(wb_path, read_only=True, data_only=False)
sheet_form = wb_form['Breakdown']

print("Loading workbook for values (read_only)...", flush=True)
wb_val = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
sheet_val = wb_val['Breakdown']

print("Extracting rows 14 to 198...", flush=True)

# Helper to get rows in read_only mode
# Since sheet.iter_rows is not optimized or doesn't support direct slicing in the same way, we can just loop and get rows
rows_form = list(sheet_form.iter_rows(min_row=11, max_row=198))
rows_val = list(sheet_val.iter_rows(min_row=11, max_row=198))

# Header is row 11 (index 0 in our lists since we started at min_row=11)
headers = [cell.value for cell in rows_val[0]]

print("\n--- Headers list ---")
for idx, h in enumerate(headers):
    if h:
        print(f"Col {idx+1} ({openpyxl.utils.get_column_letter(idx+1)}): {h}")

# Print row data for rows 14 to 198 (which is index r - 11 in our lists)
print("\n--- Row Data & Formulas ---", flush=True)
for r in range(14, 199):
    idx = r - 11
    if idx >= len(rows_form):
        break
    cells_form = rows_form[idx]
    cells_val = rows_val[idx]
    
    # Check if row is empty
    vals = [c.value for c in cells_val]
    if all(v is None for v in vals):
        continue
        
    row_type = cells_val[2].value # Col C
    komponen = cells_val[7].value # Col H
    
    print(f"\n=================== ROW {r} [Type: {row_type}] ===================")
    print(f"Komponen (Col H): {komponen}")
    
    # Print key columns
    cols_to_show = [
        ('A', 0, 'id'),
        ('B', 1, 'cat.'),
        ('C', 2, 'type'),
        ('D', 3, 'kode'),
        ('E', 4, 'Tpk'),
        ('F', 5, 'Opt'),
        ('G', 6, 'No'),
        ('H', 7, 'Komponen'),
        ('J', 9, 'P'),
        ('L', 11, 'L'),
        ('N', 13, 'T'),
        ('O', 14, 'Ukuran'),
        ('P', 15, 'Sub'),
        ('Q', 16, 'Jml'),
        ('R', 17, 'Bahan'),
        ('S', 18, 'T Bahan'),
        ('V', 21, 'Luar'),
        ('X', 23, 'Dalam'),
        ('Z', 25, 'P1'),
        ('AA', 26, 'P2'),
        ('AB', 27, 'L1'),
        ('AC', 28, 'L2'),
        ('AD', 29, 'P1 (edg)'),
        ('AE', 30, 'P2 (edg)'),
        ('AF', 31, 'L1 (edg)'),
        ('AG', 32, 'L2 (edg)'),
        ('AW', 48, 'V lap'),
        ('AX', 49, 'V edg'),
        ('AY', 50, 'Deskripsi lapisan'),
        ('AZ', 51, 'Deskripsi edging'),
        ('BA', 52, 'Deskripsi Komponen'),
        ('BB', 53, 'Nama Komponen'),
        ('BK', 62, 'P aktual'),
        ('BL', 63, 'L aktual'),
        ('BO', 66, 'Bahan Dasar'),
        ('BQ', 68, 'Harga/panel'),
        ('BW', 74, 'M2'),
        ('BX', 75, 'M3'),
        ('CF', 83, 'CSV Format')
    ]
    
    for letter, c_idx, label in cols_to_show:
        if c_idx < len(cells_form):
            c_form = cells_form[c_idx].value
            c_val = cells_val[c_idx].value
            if c_form is not None or c_val is not None:
                # Format to show formula if it starts with '='
                is_formula = str(c_form).startswith('=')
                formula_str = f" | Formula: {c_form}" if is_formula else ""
                print(f"Col {letter} ({label}): Value = {c_val}{formula_str}")

# Close workbooks
wb_form.close()
wb_val.close()
