import openpyxl

wb = openpyxl.load_workbook('/Applications/Arenga/vscode/breakdown_app/master rekap 2026_Bom.xlsx', data_only=True)
sheet = wb['Spek']
print(f"Spek Sheet Rows: {sheet.max_row}, Cols: {sheet.max_column}")
for r in range(1, min(sheet.max_row + 1, 150)):
    row_vals = [sheet.cell(r, c).value for c in range(1, sheet.max_column + 1)]
    # Filter empty rows
    if any(v is not None for v in row_vals):
        # Print row index and non-empty values
        non_empty = [(c+1, v) for c, v in enumerate(row_vals) if v is not None]
        print(f"Row {r}: {row_vals}")
