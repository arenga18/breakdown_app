import openpyxl
import json

wb_path = '/Applications/Arenga/vscode/breakdown_app/master rekap 2026_Bom.xlsx'
print("Loading workbook in read-only mode...")
wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
sheet = wb['Data Validation']

print("Workbook loaded. Extracting ranges...")

# 1. descf range: B418:C503
descf = []
for row in sheet.iter_rows(min_row=418, max_row=503, min_col=2, max_col=3, values_only=True):
    code, desc = row
    if code is not None or desc is not None:
        descf.append({
            'code': str(code).strip() if code is not None else '',
            'name': str(desc).strip() if desc is not None else ''
        })

# 2. desce range: B543:C744
desce = []
for row in sheet.iter_rows(min_row=543, max_row=744, min_col=2, max_col=3, values_only=True):
    code, desc = row
    if code is not None or desc is not None:
        desce.append({
            'code': str(code).strip() if code is not None else '',
            'name': str(desc).strip() if desc is not None else ''
        })

# 3. descfr range: B844:C866
descfr = []
for row in sheet.iter_rows(min_row=844, max_row=866, min_col=2, max_col=3, values_only=True):
    code, desc = row
    if code is not None or desc is not None:
        descfr.append({
            'code': str(code).strip() if code is not None else '',
            'name': str(desc).strip() if desc is not None else ''
        })

print(f"descf count: {len(descf)}")
print(f"desce count: {len(desce)}")
print(f"descfr count: {len(descfr)}")

# Preview items
if len(descf) > 0:
    print(f"descf first: {descf[0]}")
if len(desce) > 0:
    print(f"desce first: {desce[0]}")
if len(descfr) > 0:
    print(f"descfr first: {descfr[0]}")

with open('/Applications/Arenga/vscode/breakdown_app/scratch/extracted_desc.json', 'w') as f:
    json.dump({
        'descf': descf,
        'desce': desce,
        'descfr': descfr
    }, f, indent=2)

print("Extraction completed successfully!")
